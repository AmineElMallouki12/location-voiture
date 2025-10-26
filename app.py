from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, abort, Response
from flask_pymongo import PyMongo
from flask_bcrypt import Bcrypt
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from datetime import datetime
import json
from werkzeug.utils import secure_filename
import os
from bson.objectid import ObjectId
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from flask import send_file
import uuid

app = Flask(__name__)
app.secret_key = 'your-secret-key'

# MongoDB configuration
app.config["MONGO_URI"] = "mongodb://localhost:27017/voiture_de_location"
mongo = PyMongo(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# MySQL Configuration
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'inventory_db'
app.config['UPLOAD_FOLDER'] = os.path.join('static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024  # 2MB max upload size
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

# User Loader
class User(UserMixin):
    def __init__(self, id, username, role):
        self.id = id
        self.username = username
        self.role = role

@login_manager.user_loader
def load_user(user_id):
    user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
    if user:
        return User(str(user['_id']), user['username'], user['role'])
    return None

# Helper function to get database cursor
def get_cursor():
    return mongo.db.cars.find()

# Helper function to format datetime
def format_datetime(dt):
    if dt:
        return dt.strftime('%Y-%m-%d %H:%M')
    return None

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Login Route
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = mongo.db.users.find_one({'username': username})
        
        if user and bcrypt.check_password_hash(user['password'], password):
            login_user(User(str(user['_id']), username, user['role']))
            flash(f'Welcome back, {username}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

# Logout Route
@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out successfully', 'info')
    return redirect(url_for('login'))

# Profile Route
@app.route('/profile')
@login_required
def profile():
    user = mongo.db.users.find_one({'_id': ObjectId(current_user.id)})
    if not user:
        flash('Utilisateur introuvable', 'error')
        return redirect(url_for('dashboard'))
    # Minimal profile info
    info = {
        'username': user.get('username', ''),
        'role': user.get('role', ''),
        'first_name': user.get('first_name', ''),
        'last_name': user.get('last_name', '')
    }
    return render_template('profile.html', user=info)

# Dashboard Route
@app.route('/dashboard')
@login_required
def dashboard():
    # Get statistics
    available = mongo.db.cars.count_documents({'status': 'Disponible'})
    unavailable = mongo.db.cars.count_documents({'status': 'Indisponible'})
    repair = mongo.db.cars.count_documents({'status': 'Nécessite une réparation'})
    total = mongo.db.cars.count_documents({})
    # Get recent items
    recent_items = list(mongo.db.cars.find().sort([('updated_at', -1), ('created_at', -1)]).limit(10))
    for item in recent_items:
        # Don't overwrite the actual 'id' field with MongoDB _id
        # The 'id' field should remain as the car's custom ID
        item['created_at'] = item.get('created_at', '').strftime('%Y-%m-%d %H:%M') if item.get('created_at') else ''
        item['updated_at'] = item.get('updated_at', '').strftime('%Y-%m-%d %H:%M') if item.get('updated_at') else ''
    stats = {
        'available': available,
        'unavailable': unavailable,
        'repair': repair,
        'total': total
    }
    return render_template('dashboard.html', stats=stats, recent_items=recent_items)

# Inventory Route
@app.route('/inventory')
@login_required
def inventory():
    # Get all cars and group by category
    items = list(mongo.db.cars.find().sort('designation', 1))
    
    # Group items by category
    categorized_items = {}
    category_totals = {}
    
    for item in items:
        # Don't overwrite the actual 'id' field with MongoDB _id
        # The 'id' field should remain as the car's custom ID
        item['designation'] = item.get('designation', '')
        item['created_at'] = item.get('created_at', '').strftime('%Y-%m-%d %H:%M') if item.get('created_at') else ''
        item['updated_at'] = item.get('updated_at', '').strftime('%Y-%m-%d %H:%M') if item.get('updated_at') else ''
        
        # Get category (default to 'Non catégorisé' if not found)
        category = item.get('category', 'Non catégorisé')
        
        if category not in categorized_items:
            categorized_items[category] = []
            category_totals[category] = {
                'total_quantity': 0,
                'available_quantity': 0,
                'broken_quantity': 0,
                'repair_quantity': 0
            }
        
        categorized_items[category].append(item)
    
    return render_template('inventory.html', categorized_items=categorized_items, category_totals=category_totals)

# Add Item Route
@app.route('/add-item', methods=['GET', 'POST'])
@login_required
def add_item():
    if not (current_user.role == 'admin' or is_manager()):
        flash('Accès refusé. Réservé au manager ou admin.', 'error')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        item_id = request.form['id']
        quantity = int(request.form.get('quantity', 1))
        prix_journalier = float(request.form.get('prix_journalier', 0)) if request.form.get('prix_journalier') else 0
        carburant = request.form.get('carburant', 'Essence')
        transmission = request.form.get('transmission', 'Manuelle')
        designation = request.form['designation']
        category = request.form['category']
        marque = request.form['marque']
        modele = request.form['modele']
        n_serie = request.form['n_serie']
        ancien_cab = request.form['ancien_cab']
        nouveau_cab = request.form['nouveau_cab']
        status = request.form['status']
        date_inv = request.form['date_inv']
        description = request.form['description']
        

        
        image_file = request.files.get('image')
        image_filename = None
        if image_file and image_file.filename:
            if not allowed_file(image_file.filename):
                flash('Type de fichier image invalide. Seuls JPG, PNG, GIF sont autorisés.', 'danger')
                return redirect(request.url)
            if len(image_file.read()) > app.config['MAX_CONTENT_LENGTH']:
                flash('L\'image est trop volumineuse (max 2MB).', 'danger')
                return redirect(request.url)
            image_file.seek(0)
            filename = secure_filename(image_file.filename)
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            image_file.save(image_path)
            image_filename = filename
        now = datetime.now()
        mongo.db.cars.insert_one({
            'id': item_id,
            'quantity': quantity,
            'prix_journalier': prix_journalier,
            'carburant': carburant,
            'transmission': transmission,
            'designation': designation,
            'category': category,
            'marque': marque,
            'modele': modele,
            'n_serie': n_serie,
            'ancien_cab': ancien_cab,
            'nouveau_cab': nouveau_cab,
            'status': status,
            'date_inv': date_inv,
            'description': description,
            'quantite_totale': quantity,
            'quantite_cassée': 0,
            'quantite_en_réparation': 0,
            'quantite_disponible': quantity,
            'image': image_filename,
            'created_at': now,
            'updated_at': now
        })
        flash('Voiture ajoutée avec succès!', 'success')
        return redirect(url_for('dashboard'))
    return render_template('add_item.html')

# API Routes for AJAX requests
@app.route('/api/item/<string:item_id>', methods=['GET'])
@login_required
def get_item(item_id):
    item = mongo.db.cars.find_one({'id': item_id})
    if item:
        return jsonify({
            'success': True,
            'item': {
                'id': item.get('id', ''),
                'designation': item.get('designation', ''),
                'marque': item.get('marque', ''),
                'modele': item.get('modele', ''),
                'n_serie': item.get('n_serie', ''),
                'ancien_cab': item.get('ancien_cab', ''),
                'nouveau_cab': item.get('nouveau_cab', ''),
                'status': item.get('status', ''),
                'date_inv': item.get('date_inv', ''),
                'description': item.get('description', ''),
                'quantite_totale': item.get('quantite_totale', 1),
                'quantite_cassée': item.get('quantite_cassée', 0),
                'quantite_en_réparation': item.get('quantite_en_réparation', 0),
                'quantite_disponible': item.get('quantite_disponible', 1),
                'image': item.get('image', ''),
                'created_at': item.get('created_at', '').strftime('%Y-%m-%d %H:%M') if item.get('created_at') else '',
                'updated_at': item.get('updated_at', '').strftime('%Y-%m-%d %H:%M') if item.get('updated_at') else ''
            }
        })
    else:
        return jsonify({'success': False, 'message': 'Élément non trouvé'})

@app.route('/api/item/<string:item_id>', methods=['PUT'])
@login_required
def update_item(item_id):
    item_id_val = request.form.get('id')
    designation = request.form.get('designation')
    marque = request.form.get('marque')
    modele = request.form.get('modele')
    n_serie = request.form.get('n_serie')
    ancien_cab = request.form.get('ancien_cab')
    nouveau_cab = request.form.get('nouveau_cab')
    status = request.form.get('status')
    date_inv = request.form.get('date_inv')
    description = request.form.get('description')
    

    
    image_file = request.files.get('image')
    image_filename = None
    if image_file and image_file.filename:
        if not allowed_file(image_file.filename):
            return jsonify({'success': False, 'message': 'Type de fichier image invalide. Seuls JPG, PNG, GIF sont autorisés.'}), 400
        if len(image_file.read()) > app.config['MAX_CONTENT_LENGTH']:
            return jsonify({'success': False, 'message': 'L\'image est trop volumineuse (max 2MB).'}), 400
        image_file.seek(0)
        filename = secure_filename(image_file.filename)
        image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        image_file.save(image_path)
        image_filename = filename
    update_fields = {
        'id': item_id_val,
        'designation': designation,
        'marque': marque,
        'modele': modele,
        'n_serie': n_serie,
        'ancien_cab': ancien_cab,
        'nouveau_cab': nouveau_cab,
        'status': status,
        'date_inv': date_inv,
        'description': description,
        'quantite_totale': 1,
        'quantite_cassée': 0,
        'quantite_en_réparation': 0,
        'quantite_disponible': 1,
        'updated_at': datetime.now()
    }
    if image_filename:
        update_fields['image'] = image_filename
    mongo.db.cars.update_one({'id': item_id}, {'$set': update_fields})
    return jsonify({'success': True, 'message': 'Matériel mis à jour avec succès'})

@app.route('/api/item/<string:item_id>', methods=['DELETE'])
@login_required
def delete_item(item_id):
    mongo.db.cars.delete_one({'id': item_id})
    return jsonify({'success': True, 'message': 'Item deleted successfully'})

# Categories API
@app.route('/api/categories')
@login_required
def get_categories():
    raw_categories = mongo.db.cars.distinct('category')
    categories = [c for c in raw_categories if c and str(c).strip()]
    categories.sort(key=lambda x: str(x).lower())
    return jsonify({'success': True, 'categories': categories})

# Available items API (optionally filtered by category)
@app.route('/api/available-items')
@login_required
def get_available_items():
    selected_category = (request.args.get('category') or '').strip()

    query = {
        'status': 'Disponible',
        'quantite_disponible': {'$gt': 0}
    }
    if selected_category:
        query['category'] = selected_category

    available_items = list(
        mongo.db.cars.find(query).sort('designation', 1)
    )

    items = []
    for item in available_items:
        items.append({
            'id': item.get('id', ''),
            'designation': item.get('designation', ''),
            'condition': item.get('condition', ''),
            'quantite_disponible': item.get('quantite_disponible', 0)
        })

    return jsonify({'success': True, 'items': items})

# Reservation API routes
@app.route('/api/reservation', methods=['POST'])
@login_required
def create_reservation():
    # Only utilisateur and manager can create reservations
    if not (is_utilisateur() or is_manager()):
        return jsonify({'success': False, 'message': 'Accès refusé. Réservé aux utilisateurs et managers.'})
    
    item_id = request.form.get('item_id')
    user_name = request.form.get('user_name')
    quantity = int(request.form.get('quantity', 1))
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    purpose = request.form.get('purpose', '')
    
    if not all([item_id, user_name, start_date, end_date]):
        return jsonify({'success': False, 'message': 'Tous les champs requis doivent être fournis'})
    
    # Check if item exists and is available
    item = mongo.db.cars.find_one({'id': item_id})
    if not item:
        return jsonify({'success': False, 'message': 'Équipement non trouvé'})
    
    available_quantity = item.get('quantite_disponible', item.get('quantite_totale', 1))
    if item.get('status') != 'Disponible' or available_quantity <= 0:
        return jsonify({'success': False, 'message': 'L\'équipement n\'est pas disponible pour la réservation'})
    
    # Check if requested quantity is available
    if quantity > available_quantity:
        return jsonify({'success': False, 'message': f'Seulement {available_quantity} unités disponibles. Vous avez demandé {quantity}.'})
    
    # Create reservation
    reservation_data = {
        'item_id': item_id,
        'user_name': user_name,
        'user_email': current_user.username,  # Use current user's username as email
        'quantity': quantity,
        'start_date': datetime.strptime(start_date, '%Y-%m-%dT%H:%M'),
        'end_date': datetime.strptime(end_date, '%Y-%m-%dT%H:%M'),
        'purpose': purpose,
        'status': 'Pending',
        'created_at': datetime.now()
    }
    
    mongo.db.rental_requests.insert_one(reservation_data)
    
    return jsonify({'success': True, 'message': 'Demande de réservation créée avec succès'})

@app.route('/api/reservation/<string:reservation_id>/approve', methods=['PUT'])
@login_required
def approve_reservation(reservation_id):
    # Check if user is manager or admin
    if not (is_manager() or current_user.role == 'admin'):
        return jsonify({'success': False, 'message': 'Accès refusé. Réservé au manager ou admin.'}), 403
    
    try:
        # Validate ObjectId format
        try:
            object_id = ObjectId(reservation_id)
        except Exception as e:
            return jsonify({'success': False, 'message': f'ID de réservation invalide: {reservation_id}'}), 400
        
        # Find the reservation
        reservation = mongo.db.rental_requests.find_one({'_id': object_id})
        if not reservation:
            return jsonify({'success': False, 'message': 'Réservation non trouvée'}), 404
        
        # Check if reservation is pending
        current_status = reservation.get('status')
        if current_status != 'En attente':
            return jsonify({'success': False, 'message': 'Seules les réservations en attente peuvent être approuvées'}), 400
        
        # Update reservation status
        mongo.db.rental_requests.update_one(
            {'_id': object_id},
            {'$set': {
                'status': 'Approved', 
                'approved_by': f'{current_user.role}:{current_user.username}', 
                'approved_at': datetime.now()
            }}
        )
        
        flash(f'Réservation approuvée avec succès!', 'success')
        return jsonify({'success': True, 'message': f'Réservation approuvée avec succès'})
        
    except Exception as e:
        print(f"Error approving reservation: {e}")
        return jsonify({'success': False, 'message': f'Erreur lors de l\'approbation: {str(e)}'}), 500

@app.route('/api/reservation/<string:reservation_id>/reject', methods=['PUT'])
@login_required
def reject_reservation(reservation_id):
    # Check if user is manager or admin
    if not (is_manager() or current_user.role == 'admin'):
        return jsonify({'success': False, 'message': 'Accès refusé. Réservé au manager ou admin.'}), 403
    
    try:
        # Validate ObjectId format
        try:
            object_id = ObjectId(reservation_id)
        except Exception as e:
            return jsonify({'success': False, 'message': f'ID de réservation invalide: {reservation_id}'}), 400
        
        # Find the reservation
        reservation = mongo.db.rental_requests.find_one({'_id': object_id})
        if not reservation:
            return jsonify({'success': False, 'message': 'Réservation non trouvée'}), 404
        
        # Check if reservation can be rejected
        current_status = reservation.get('status')
        if current_status != 'En attente':
            return jsonify({'success': False, 'message': 'Seules les réservations en attente peuvent être rejetées'}), 400
        
        # Update status to rejected
        mongo.db.rental_requests.update_one(
            {'_id': object_id},
            {'$set': {
                'status': 'Rejected',
                'rejected_by': current_user.username,
                'rejected_at': datetime.now()
            }}
        )
        
        # Restore quantities for all items in the reservation
        if 'items' in reservation:
            # Multi-item reservation
            for item_data in reservation['items']:
                item_id = item_data.get('item_id')
                quantity = item_data.get('quantity', 1)
                
                # Restore the quantity
                mongo.db.cars.update_one(
                    {'id': item_id},
                    {
                        '$inc': {'quantite_disponible': quantity},
                        '$set': {
                            'status': 'Disponible',
                            'updated_at': datetime.now()
                        }
                    }
                )
        else:
            # Single item reservation (legacy)
            item_id = reservation.get('item_id')
            quantity = reservation.get('quantity', 1)
            
            if item_id:
                mongo.db.cars.update_one(
                    {'id': item_id},
                    {
                        '$inc': {'quantite_disponible': quantity},
                        '$set': {
                            'status': 'Disponible',
                            'updated_at': datetime.now()
                        }
                    }
                )
        
        flash('Réservation rejetée avec succès!', 'success')
        return jsonify({'success': True, 'message': 'Réservation rejetée avec succès'})
        
    except Exception as e:
        print(f"Error rejecting reservation: {e}")
        return jsonify({'success': False, 'message': f'Erreur lors du rejet: {str(e)}'}), 500

@app.route('/api/reservation/<string:reservation_id>', methods=['DELETE'])
@login_required
def delete_reservation(reservation_id):
    """Delete a reservation"""
    try:
        reservation = mongo.db.rental_requests.find_one({'_id': ObjectId(reservation_id)})
        if not reservation:
            return jsonify({'success': False, 'message': 'Réservation non trouvée'}), 404
        
        # Check if user can delete this reservation
        if not (is_manager() or current_user.role == 'admin' or reservation.get('user_email') == current_user.username):
            return jsonify({'success': False, 'message': 'Permission refusée'}), 403
        
        # If reservation is pending, restore quantities
        if reservation.get('status') == 'En attente':
            if 'items' in reservation and reservation['items']:
                # Multi-item reservation
                for item_data in reservation['items']:
                    item_id = item_data.get('item_id')
                    quantity = item_data.get('quantity', 1)
                    mongo.db.cars.update_one(
                        {'id': item_id},
                        {'$inc': {'quantite_disponible': quantity}, '$set': {'status': 'Disponible', 'updated_at': datetime.now()}}
                    )
            else:
                # Single item reservation (legacy)
                item_id = reservation.get('item_id')
                quantity = reservation.get('quantity', 1)
                if item_id:
                    mongo.db.cars.update_one(
                        {'id': item_id},
                        {'$inc': {'quantite_disponible': quantity}, '$set': {'status': 'Disponible', 'updated_at': datetime.now()}}
                    )
        
        # Delete the reservation
        mongo.db.rental_requests.delete_one({'_id': ObjectId(reservation_id)})
        
        return jsonify({'success': True, 'message': 'Réservation supprimée avec succès'})
        
    except Exception as e:
        print(f"Error deleting reservation: {e}")
        return jsonify({'success': False, 'message': 'Erreur lors de la suppression'}), 500

@app.route('/api/reservation/<string:reservation_id>', methods=['GET'])
@login_required
def get_reservation_details(reservation_id):
    """Return detailed information about a reservation, including car name and category."""
    try:
        reservation = mongo.db.rental_requests.find_one({'_id': ObjectId(reservation_id)})
    except Exception:
        reservation = None
    if not reservation:
        return jsonify({'success': False, 'message': 'Réservation non trouvée'}), 404

    # Authorization: allow admin/manager, the owner
    is_owner = reservation.get('user_email') == current_user.username
    can_view = is_owner or is_manager() or current_user.role == 'admin'
    if not can_view:
        return jsonify({'success': False, 'message': 'Accès refusé'}), 403

    # Handle multi-item reservations
    if 'items' in reservation:
        # Multi-item reservation
        items_data = []
        for item_data in reservation['items']:
            item_id = item_data.get('item_id')
            quantity = item_data.get('quantity', 1)
            designation = item_data.get('designation', '')
            
            car = mongo.db.cars.find_one({'id': item_id}) or {}
            items_data.append({
                'item_id': item_id,
                'designation': designation or car.get('designation', ''),
                'category': car.get('category', ''),
                'quantity': quantity
            })
        
        data = {
            'id': str(reservation.get('_id')),
            'user_name': reservation.get('user_name', ''),
            'user_email': reservation.get('user_email', ''),
            'start_date': reservation.get('start_date').strftime('%Y-%m-%d %H:%M') if reservation.get('start_date') else '',
            'end_date': reservation.get('end_date').strftime('%Y-%m-%d %H:%M') if reservation.get('end_date') else '',
            'status': reservation.get('status', ''),
            'purpose': reservation.get('purpose', ''),
            'is_multi_item': True,
            'items': items_data,
            'total_items': len(items_data)
        }
    else:
        # Single item reservation (legacy)
        car = mongo.db.cars.find_one({'id': reservation.get('item_id')}) or {}
        data = {
            'id': str(reservation.get('_id')),
            'item_id': reservation.get('item_id', ''),
            'item_name': car.get('designation', ''),
            'category': car.get('category', ''),
            'user_name': reservation.get('user_name', ''),
            'user_email': reservation.get('user_email', ''),
            'quantity': reservation.get('quantity', 1),
            'start_date': reservation.get('start_date').strftime('%Y-%m-%d %H:%M') if reservation.get('start_date') else '',
            'end_date': reservation.get('end_date').strftime('%Y-%m-%d %H:%M') if reservation.get('end_date') else '',
            'status': reservation.get('status', ''),
            'purpose': reservation.get('purpose', ''),
            'is_multi_item': False
        }
    
    return jsonify({'success': True, 'reservation': data})

@app.route('/api/create-reservation', methods=['POST'])
@login_required
def create_cart_reservation():
    """Create a new reservation from shopping cart with multiple items"""
    try:
        data = request.get_json()
        
        if not data or 'items' not in data:
            return jsonify({'success': False, 'message': 'Données invalides'}), 400
        
        items = data.get('items', [])
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        purpose = data.get('purpose')
        
        if not items or not start_date or not end_date or not purpose:
            return jsonify({'success': False, 'message': 'Tous les champs sont obligatoires'}), 400
        
        # Validate dates
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%dT%H:%M')
            end_dt = datetime.strptime(end_date, '%Y-%m-%dT%H:%M')
            if start_dt >= end_dt:
                return jsonify({'success': False, 'message': 'La date de fin doit être postérieure à la date de début'}), 400
        except ValueError:
            return jsonify({'success': False, 'message': 'Format de date invalide'}), 400
        
        # Check if all items are available
        for item_data in items:
            item_id = item_data.get('item_id')
            quantity = item_data.get('quantity', 1)
            
            if not item_id or quantity <= 0:
                return jsonify({'success': False, 'message': 'Données d\'article invalides'}), 400
            
            # Check if item exists and has enough quantity
            item = mongo.db.cars.find_one({'id': item_id})
            if not item:
                return jsonify({'success': False, 'message': f'Article {item_id} non trouvé'}), 400
            
            if item.get('quantite_disponible', 0) < quantity:
                article_name = item.get("designation", "l'article")
                return jsonify({'success': False, 'message': f'Quantité insuffisante pour {article_name}'}), 400
        
        # Create ONE reservation with multiple items
        reservation_data = {
            'user_name': current_user.username,
            'user_email': current_user.username,
            'start_date': start_dt,
            'end_date': end_dt,
            'purpose': purpose,
            'status': 'En attente',
            'created_at': datetime.now(),
            'items': []  # Array to store all items in this reservation
        }
        
        # Add all items to the reservation
        for item_data in items:
            item_id = item_data.get('item_id')
            quantity = item_data.get('quantity', 1)
            
            item = mongo.db.cars.find_one({'id': item_id})
            
            # Add item to reservation
            reservation_data['items'].append({
                'item_id': item_id,
                'designation': item.get('designation', ''),
                'quantity': quantity
            })
            
            # Update item availability
            new_available = item.get('quantite_disponible', 0) - quantity
            
            # Update status based on remaining quantity
            if new_available <= 0:
                new_status = "Indisponible"
            else:
                new_status = "Disponible"
            
            mongo.db.cars.update_one(
                {'id': item_id},
                {
                    '$set': {
                        'quantite_disponible': new_available,
                        'status': new_status,
                        'updated_at': datetime.now()
                    }
                }
            )
        
        # Insert the single reservation
        result = mongo.db.rental_requests.insert_one(reservation_data)
        
        return jsonify({
            'success': True, 
            'message': f'Réservation créée avec succès ({len(items)} articles)',
            'reservation_id': str(result.inserted_id)
        })
        
    except Exception as e:
        print(f"Error creating cart reservation: {e}")
        return jsonify({'success': False, 'message': 'Erreur lors de la création de la réservation'}), 500

# Reservations Route
@app.route('/reservations')
@login_required
def reservations():
    # Get reservations based on user role, excluding rejected and completed ones
    base_filter = {'status': {'$nin': ['Rejected', 'Completed']}}
    
    if is_manager() or current_user.role == 'admin':
        # Manager and admin see all active reservations (not rejected/completed)
        reservations = list(mongo.db.rental_requests.find(base_filter).sort('created_at', -1))
    elif is_utilisateur():
        # Utilisateur sees only their own reservations (not rejected/completed)
        reservations = list(mongo.db.rental_requests.find({
            '$and': [
                base_filter,
                {'user_email': current_user.username}
            ]
        }).sort('created_at', -1))
    else:
        # Default: user sees only their own reservations (not rejected/completed)
        user_filter = {'$and': [base_filter, {'user_email': current_user.username}]}
        reservations = list(mongo.db.rental_requests.find(user_filter).sort('created_at', -1))
    
    # Build a lookup for car names keyed by custom 'id'
    cars_map = {e.get('id', ''): e.get('designation', '') for e in mongo.db.cars.find()}
    
    # Format reservations for template
    formatted_reservations = []
    for r in reservations:
        if 'items' in r and r['items']:
            # Multi-item reservation
            item_names = []
            total_quantity = 0
            for item_data in r['items']:
                item_id = item_data.get('item_id', '')
                quantity = item_data.get('quantity', 1)
                designation = item_data.get('designation', '') or cars_map.get(item_id, '')
                item_names.append(f"{designation} (x{quantity})")
                total_quantity += quantity
            
            formatted_reservations.append({
                'id': str(r.get('_id')),
                'item_id': 'multi',  # Mark as multi-item
                'item_name': ' + '.join(item_names),
                'user_name': r.get('user_name', ''),
                'user_email': r.get('user_email', ''),
                'quantity': total_quantity,
                'start_date': r.get('start_date', '').strftime('%Y-%m-%d %H:%M') if r.get('start_date') else '',
                'end_date': r.get('end_date', '').strftime('%Y-%m-%d %H:%M') if r.get('end_date') else '',
                'status': r.get('status', ''),
                'purpose': r.get('purpose', ''),
                'created_at': r.get('created_at', '').strftime('%Y-%m-%d %H:%M') if r.get('created_at') else '',
                'is_multi_item': True
            })
        else:
            # Single item reservation (legacy)
            formatted_reservations.append({
                'id': str(r.get('_id')),
                'item_id': str(r.get('item_id')),
                'item_name': cars_map.get(str(r.get('item_id')), ''),
                'user_name': r.get('user_name', ''),
                'user_email': r.get('user_email', ''),
                'quantity': r.get('quantity', 1),
                'start_date': r.get('start_date', '').strftime('%Y-%m-%d %H:%M') if r.get('start_date') else '',
                'end_date': r.get('end_date', '').strftime('%Y-%m-%d %H:%M') if r.get('end_date') else '',
                'status': r.get('status', ''),
                'purpose': r.get('purpose', ''),
                'created_at': r.get('created_at', '').strftime('%Y-%m-%d %H:%M') if r.get('created_at') else '',
                'is_multi_item': False
            })
    
    return render_template('reservation.html', reservations=formatted_reservations)

# API route to get full reservation history (including rejected/completed)
@app.route('/api/reservations/history')
@login_required
def get_reservation_history():
    """Get all reservations including rejected and completed ones for history view"""
    try:
        # Get reservations based on user role
        if is_manager() or current_user.role == 'admin':
            # Manager and admin see all reservations
            reservations = list(mongo.db.rental_requests.find().sort('created_at', -1))
        else:
            # Utilisateur sees only their own reservations
            reservations = list(mongo.db.rental_requests.find({'user_email': current_user.username}).sort('created_at', -1))
        
        # Build a lookup for car names keyed by custom 'id'
        cars_map = {e.get('id', ''): e.get('designation', '') for e in mongo.db.cars.find()}
        
        # Format reservations for API response
        formatted_reservations = []
        for r in reservations:
            if 'items' in r and r['items']:
                # Multi-item reservation
                item_names = []
                total_quantity = 0
                for item_data in r['items']:
                    item_id = item_data.get('item_id', '')
                    quantity = item_data.get('quantity', 1)
                    designation = item_data.get('designation', '') or cars_map.get(item_id, '')
                    item_names.append(f"{designation} (x{quantity})")
                    total_quantity += quantity
                
                formatted_reservations.append({
                    'id': str(r.get('_id')),
                    'item_id': 'multi',
                    'item_name': ' + '.join(item_names),
                    'user_name': r.get('user_name', ''),
                    'user_email': r.get('user_email', ''),
                    'quantity': total_quantity,
                    'start_date': r.get('start_date', '').strftime('%Y-%m-%d %H:%M') if r.get('start_date') else '',
                    'end_date': r.get('end_date', '').strftime('%Y-%m-%d %H:%M') if r.get('end_date') else '',
                    'status': r.get('status', ''),
                    'purpose': r.get('purpose', ''),
                    'created_at': r.get('created_at', '').strftime('%Y-%m-%d %H:%M') if r.get('created_at') else '',
                    'is_multi_item': True
                })
            else:
                # Single item reservation (legacy)
                formatted_reservations.append({
                    'id': str(r.get('_id')),
                    'item_id': str(r.get('item_id')),
                    'item_name': cars_map.get(str(r.get('item_id')), ''),
                    'user_name': r.get('user_name', ''),
                    'user_email': r.get('user_email', ''),
                    'quantity': r.get('quantity', 1),
                    'start_date': r.get('start_date', '').strftime('%Y-%m-%d %H:%M') if r.get('start_date') else '',
                    'end_date': r.get('end_date', '').strftime('%Y-%m-%d %H:%M') if r.get('end_date') else '',
                    'status': r.get('status', ''),
                    'purpose': r.get('purpose', ''),
                    'created_at': r.get('created_at', '').strftime('%Y-%m-%d %H:%M') if r.get('created_at') else '',
                    'is_multi_item': False
                })
        
        return jsonify({
            'success': True,
            'reservations': formatted_reservations
        })
        
    except Exception as e:
        print(f"Error getting reservation history: {e}")
        return jsonify({
            'success': False,
            'message': f'Erreur lors de la récupération de l\'historique: {str(e)}'
        }), 500

# Report Generation Route
@app.route('/generate-report', methods=['GET', 'POST'])
@login_required
def generate_report():
    # Students cannot access reports
    if current_user.role == 'etudiant':
        flash('Accès refusé. Les étudiants ne peuvent pas accéder aux rapports.', 'error')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        report_type = request.form.get('report_type', 'inventory')

        if report_type == 'inventory':
            items = []
            for item in mongo.db.cars.find().sort('designation', 1):
                items.append({
                    'id': item.get('id', ''),
                    'designation': item.get('designation', ''),
                    'marque': item.get('marque', ''),
                    'modele': item.get('modele', ''),
                    'n_serie': item.get('n_serie', ''),
                    'ancien_cab': item.get('ancien_cab', ''),
                    'nouveau_cab': item.get('nouveau_cab', ''),
                    'status': item.get('status', ''),
                    'date_inv': item.get('date_inv', ''),
                    'description': item.get('description', '')
                })
            return render_template('report.html', items=items, report_type='inventory')

        elif report_type == 'statistics':
            available = mongo.db.cars.count_documents({'status': 'Disponible'})
            unavailable = mongo.db.cars.count_documents({'status': 'Indisponible'})
            repair = mongo.db.cars.count_documents({'status': 'Nécessite une réparation'})
            total = mongo.db.cars.count_documents({})
            # Condition repartition
            bon_etat = mongo.db.cars.count_documents({'condition': 'Bon état'})
            mauvais_etat = mongo.db.cars.count_documents({'condition': 'Mauvais état'})
            condition_repair = mongo.db.cars.count_documents({'condition': 'Nécessite une réparation'})
            autre = total - (bon_etat + mauvais_etat + condition_repair)
            stats = {
                'available': available,
                'unavailable': unavailable,
                'repair': repair,
                'total': total,
                'bon_etat': bon_etat,
                'mauvais_etat': mauvais_etat,
                'condition_repair': condition_repair,
                'autre': autre
            }
            return render_template('report.html', stats=stats, report_type='statistics')

        elif report_type == 'reservations':
            # You can add reservation report logic here if needed
            return render_template('report.html', report_type='reservations')

    return render_template('report.html')

@app.route('/export-report/pdf')
@login_required
def export_report_pdf():
    # Students cannot export reports
    if current_user.role == 'etudiant':
        flash('Accès refusé. Les étudiants ne peuvent pas exporter des rapports.', 'error')
        return redirect(url_for('dashboard'))
    report_type = request.args.get('report_type', 'inventory')
    output = BytesIO()
    c = canvas.Canvas(output, pagesize=landscape(letter))
    width, height = landscape(letter)
    y = height - 40
    c.setFont('Helvetica-Bold', 18)
    title_map = {
        'inventory': 'Rapport d\'inventaire (détaillé)',
        'statistics': 'Rapport de statistiques',
        'reservations': 'Rapport des réservations (détaillé)'
    }
    c.drawString(40, y, title_map.get(report_type, 'Rapport'))
    c.setFont('Helvetica', 12)
    y -= 40
    if report_type == 'inventory':
        items = list(mongo.db.cars.find().sort('designation', 1))
        headers = [
            'ID', 'Désignation', 'Catégorie', 'Marque', 'Modèle', 'N° Série',
            'Ancien CAB', 'Nouveau CAB', 'Statut', "Date d'inventaire",
            'Qté Tot.', 'Qté Disp.', 'Qté Cassée', 'Qté Réparation', 'Description'
        ]
        c.setFont('Helvetica-Bold', 10)
        c.drawString(40, y, ' | '.join(headers))
        y -= 18
        c.setFont('Helvetica', 9)
        for item in items:
            description = (item.get('description', '') or '').replace('\n', ' ')
            if len(description) > 80:
                description = description[:77] + '...'
            row = [
                str(item.get('id', '')),
                str(item.get('designation', '')),
                str(item.get('category', '')),
                str(item.get('marque', '')),
                str(item.get('modele', '')),
                str(item.get('n_serie', '')),
                str(item.get('ancien_cab', '')),
                str(item.get('nouveau_cab', '')),
                str(item.get('status', '')),
                str(item.get('date_inv', '')),
                str(item.get('quantite_totale', 1)),
                str(item.get('quantite_disponible', item.get('quantite_totale', 1))),
                str(item.get('quantite_cassée', 0)),
                str(item.get('quantite_en_réparation', 0)),
                description
            ]
            c.drawString(40, y, ' | '.join(row))
            y -= 16
            if y < 40:
                c.showPage()
                y = height - 40
                c.setFont('Helvetica-Bold', 10)
                c.drawString(40, y, ' | '.join(headers))
                y -= 18
                c.setFont('Helvetica', 9)
    elif report_type == 'statistics':
        available = mongo.db.cars.count_documents({'status': 'Disponible'})
        unavailable = mongo.db.cars.count_documents({'status': 'Indisponible'})
        repair = mongo.db.cars.count_documents({'status': 'Nécessite une réparation'})
        total = mongo.db.cars.count_documents({})
        # Répartition par condition si existante
        bon_etat = mongo.db.cars.count_documents({'condition': 'Bon état'})
        mauvais_etat = mongo.db.cars.count_documents({'condition': 'Mauvais état'})
        condition_repair = mongo.db.cars.count_documents({'condition': 'Nécessite une réparation'})
        autre = total - (bon_etat + mauvais_etat + condition_repair)
        stats = [
            ('Disponible', available),
            ('Indisponible', unavailable),
            ('Nécessite une réparation', repair),
            ('Total', total),
            ('Bon état', bon_etat),
            ('Mauvais état', mauvais_etat),
            ('Autre', autre)
        ]
        for label, value in stats:
            c.drawString(40, y, f"{label}: {value}")
            y -= 24
    elif report_type == 'reservations':
        reservations = list(mongo.db.rental_requests.find().sort('start_date', -1))
        car_map = {e.get('id', ''): e.get('designation', '') for e in mongo.db.cars.find()}
        headers = ['Article', 'Catégorie', 'Réservé par', 'Email', 'Qté', 'Début', 'Fin', 'Statut', 'But']
        c.drawString(40, y, ' | '.join(headers))
        y -= 20
        for r in reservations:
            if 'items' in r:
                # Multi-item request - create a row for each item
                for item_data in r.get('items', []):
                    equip = mongo.db.cars.find_one({'id': item_data.get('item_id')}) or {}
                    row = [
                        car_map.get(item_data.get('item_id', ''), ''),
                        str(equip.get('category', '')),
                        r.get('user_name', ''),
                        r.get('user_email', ''),
                        str(item_data.get('quantity', 1)),
                        r.get('start_date', '').strftime('%Y-%m-%d %H:%M') if r.get('start_date') else '',
                        r.get('end_date', '').strftime('%Y-%m-%d %H:%M') if r.get('end_date') else '',
                        r.get('status', ''),
                        (r.get('purpose', '') or '')[:60]
                    ]
                    row_str = ' | '.join(row)
                    c.drawString(40, y, row_str)
                    y -= 18
                    if y < 40:
                        c.showPage()
                        y = height - 40
            else:
                # Single-item request (legacy)
                equip = mongo.db.cars.find_one({'id': r.get('item_id')}) or {}
                row = [
                    car_map.get(r.get('item_id', ''), ''),
                    str(equip.get('category', '')),
                    r.get('user_name', ''),
                    r.get('user_email', ''),
                    str(r.get('quantity', 1)),
                    r.get('start_date', '').strftime('%Y-%m-%d %H:%M') if r.get('start_date') else '',
                    r.get('end_date', '').strftime('%Y-%m-%d %H:%M') if r.get('end_date') else '',
                    r.get('status', ''),
                    (r.get('purpose', '') or '')[:60]
                ]
                row_str = ' | '.join(row)
                c.drawString(40, y, row_str)
                y -= 18
                if y < 40:
                    c.showPage()
                    y = height - 40
    c.save()
    output.seek(0)
    name_map = {
        'inventory': 'inventaire_detaillé',
        'statistics': 'statistiques',
        'reservations': 'reservations_detaillé'
    }
    return send_file(output, as_attachment=True, download_name=f"{name_map.get(report_type, 'rapport')}.pdf", mimetype='application/pdf')

@app.route('/export-report/excel')
@login_required
def export_report_excel():
    # Students cannot export reports
    if current_user.role == 'etudiant':
        flash('Accès refusé. Les étudiants ne peuvent pas exporter des rapports.', 'error')
        return redirect(url_for('dashboard'))
    report_type = request.args.get('report_type', 'inventory')
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Rapport'
    if report_type == 'inventory':
        items = list(mongo.db.cars.find().sort('designation', 1))
        headers = [
            'ID', 'Désignation', 'Catégorie', 'Marque', 'Modèle', 'N° Série',
            'Ancien CAB', 'Nouveau CAB', 'Statut', "Date d'inventaire",
            'Qté Totale', 'Qté Disponible', 'Qté Cassée', 'Qté en Réparation', 'Description'
        ]
        ws.append(headers)
        for item in items:
            row = [
                item.get('id', ''),
                item.get('designation', ''),
                item.get('category', ''),
                item.get('marque', ''),
                item.get('modele', ''),
                item.get('n_serie', ''),
                item.get('ancien_cab', ''),
                item.get('nouveau_cab', ''),
                item.get('status', ''),
                item.get('date_inv', ''),
                item.get('quantite_totale', 1),
                item.get('quantite_disponible', item.get('quantite_totale', 1)),
                item.get('quantite_cassée', 0),
                item.get('quantite_en_réparation', 0),
                item.get('description', '')
            ]
            ws.append(row)
    elif report_type == 'statistics':
        available = mongo.db.cars.count_documents({'status': 'Disponible'})
        unavailable = mongo.db.cars.count_documents({'status': 'Indisponible'})
        repair = mongo.db.cars.count_documents({'status': 'Nécessite une réparation'})
        total = mongo.db.cars.count_documents({})
        ws.append(['Libellé', 'Valeur'])
        ws.append(['Disponible', available])
        ws.append(['Indisponible', unavailable])
        ws.append(['Nécessite une réparation', repair])
        ws.append(['Total', total])
        # Répartition par condition (si disponible)
        bon_etat = mongo.db.cars.count_documents({'condition': 'Bon état'})
        mauvais_etat = mongo.db.cars.count_documents({'condition': 'Mauvais état'})
        condition_repair = mongo.db.cars.count_documents({'condition': 'Nécessite une réparation'})
        autre = total - (bon_etat + mauvais_etat + condition_repair)
        ws.append([])
        ws.append(['Condition', 'Valeur'])
        ws.append(['Bon état', bon_etat])
        ws.append(['Mauvais état', mauvais_etat])
        ws.append(['Nécessite une réparation', condition_repair])
        ws.append(['Autre', autre])
    elif report_type == 'reservations':
        reservations = list(mongo.db.rental_requests.find().sort('start_date', -1))
        car_map = {e.get('id', ''): e.get('designation', '') for e in mongo.db.cars.find()}
        headers = ['Article', 'Catégorie', 'Réservé par', 'Email', 'Quantité', 'Date début', 'Date fin', 'Statut', 'But']
        ws.append(headers)
        for r in reservations:
            if 'items' in r:
                # Multi-item request - create a row for each item
                for item_data in r.get('items', []):
                    equip = mongo.db.cars.find_one({'id': item_data.get('item_id')}) or {}
                    row = [
                        car_map.get(item_data.get('item_id', ''), ''),
                        equip.get('category', ''),
                        r.get('user_name', ''),
                        r.get('user_email', ''),
                        item_data.get('quantity', 1),
                        r.get('start_date', '').strftime('%Y-%m-%d %H:%M') if r.get('start_date') else '',
                        r.get('end_date', '').strftime('%Y-%m-%d %H:%M') if r.get('end_date') else '',
                        r.get('status', ''),
                        r.get('purpose', '')
                    ]
                    ws.append(row)
            else:
                # Single-item request (legacy)
                equip = mongo.db.cars.find_one({'id': r.get('item_id')}) or {}
                row = [
                    car_map.get(r.get('item_id', ''), ''),
                    equip.get('category', ''),
                    r.get('user_name', ''),
                    r.get('user_email', ''),
                    r.get('quantity', 1),
                    r.get('start_date', '').strftime('%Y-%m-%d %H:%M') if r.get('start_date') else '',
                    r.get('end_date', '').strftime('%Y-%m-%d %H:%M') if r.get('end_date') else '',
                    r.get('status', ''),
                    r.get('purpose', '')
                ]
                ws.append(row)
    wb.save(output)
    output.seek(0)
    name_map = {
        'inventory': 'inventaire_detaillé',
        'statistics': 'statistiques',
        'reservations': 'reservations_detaillé'
    }
    return send_file(output, as_attachment=True, download_name=f"{name_map.get(report_type, 'rapport')}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Direct inventory Excel export (detailed) used by inventory page
@app.route('/export/inventory-excel')
@login_required
def export_inventory_excel():
    # Students cannot export inventory
    if current_user.role == 'etudiant':
        flash('Accès refusé. Les étudiants ne peuvent pas exporter l\'inventaire.', 'error')
        return redirect(url_for('dashboard'))
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventaire'

    headers = [
        'Catégorie', 'ID', 'Désignation', 'Marque', 'Modèle', 'N° Série',
        'Ancien CAB', 'Nouveau CAB', "Date d'inventaire", 'Quantité Totale',
        'Disponible', 'Cassée', 'En Réparation', 'Statut', 'État', 'Description / Observation'
    ]
    ws.append(headers)

    # Query items grouped by category then designation
    items = list(mongo.db.cars.find().sort([('category', 1), ('designation', 1)]))
    for it in items:
        row = [
            it.get('category', ''),
            it.get('id', ''),
            it.get('designation', ''),
            it.get('marque', ''),
            it.get('modele', ''),
            it.get('n_serie', ''),
            it.get('ancien_cab', ''),
            it.get('nouveau_cab', ''),
            it.get('date_inv', ''),
            it.get('quantite_totale', 1),
            it.get('quantite_disponible', it.get('quantite_totale', 1)),
            it.get('quantite_cassée', 0),
            it.get('quantite_en_réparation', 0),
            it.get('status', ''),
            it.get('condition', ''),
            it.get('description', '')
        ]
        ws.append(row)

    # Force text format for sensitive columns to prevent Excel auto-formatting
    sensitive_cols = ['B', 'F', 'G', 'H']  # ID, N° Série, Ancien CAB, Nouveau CAB
    for col_letter in sensitive_cols:
        for cell in ws[col_letter]:
            cell.number_format = '@'

    # Autosize columns
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 60)

    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='inventaire_detaille.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/create-default-users')
def create_default_users():
    if mongo.db.users.find_one({'username': 'admin'}):
        return Response('Admin user already exists.', mimetype='text/plain')
    admin_pw = bcrypt.generate_password_hash('admin123').decode('utf-8')
    staff_pw = bcrypt.generate_password_hash('staff123').decode('utf-8')
    mongo.db.users.insert_one({
        'username': 'admin',
        'password': admin_pw,
        'role': 'admin'
    })
    mongo.db.users.insert_one({
        'username': 'staff',
        'password': staff_pw,
        'role': 'staff'
    })
    return Response('Default admin and staff users created.\nAdmin: admin/admin123\nStaff: staff/staff123', mimetype='text/plain')

@app.route('/')
def index():
    # Public car catalog - no login required
    # Only show available car with available quantity > 0
    items = list(mongo.db.cars.find({
        '$or': [
            {'status': 'Disponible'},
            {'status': 'Available'}
        ]
    }).sort('designation', 1))
    
    # Filter items with available quantity > 0
    available_items = []
    for item in items:
        quantite_disponible = item.get('quantite_disponible', item.get('quantite_totale', 1))
        if quantite_disponible > 0:
            # Don't overwrite the actual 'id' field with MongoDB _id
            # The 'id' field should remain as the car's custom ID
            item['created_at'] = item.get('created_at', '').strftime('%Y-%m-%d %H:%M') if item.get('created_at') else ''
            item['updated_at'] = item.get('updated_at', '').strftime('%Y-%m-%d %H:%M') if item.get('updated_at') else ''
            item['quantite_disponible'] = quantite_disponible
            available_items.append(item)
    
    return render_template('public_catalog.html', items=available_items)

# Helper functions for role checks

def is_manager():
    return current_user.role == 'manager'

def is_utilisateur():
    return current_user.role == 'utilisateur'

# Reservation request route for utilisateurs and managers
@app.route('/request-rental/<string:item_id>', methods=['GET', 'POST'])
@login_required
def request_rental(item_id):
    # Only utilisateur and manager can request
    if not (is_utilisateur() or is_manager()):
        flash('Accès refusé. Réservé aux utilisateurs et managers.', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        user_name = request.form.get('user_name')
        user_email = request.form.get('user_email')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        quantity = int(request.form.get('quantity', 1))
        purpose = request.form.get('purpose', '')
        
        if not all([user_name, user_email, start_date, end_date]):
            flash('Please fill in all required fields', 'error')
            return redirect(url_for('request_rental', item_id=item_id))
        
        # Check if requested quantity is available
        item = mongo.db.cars.find_one({'id': item_id})
        if not item:
            flash('Voiture non trouvée', 'error')
            return redirect(url_for('index'))
        
        available_quantity = item.get('quantite_disponible', item.get('quantite_totale', 1))
        if quantity > available_quantity:
            flash(f'Seulement {available_quantity} unités disponibles. Vous avez demandé {quantity}.', 'error')
            return redirect(url_for('request_rental', item_id=item_id))
        
        # Create rental request
        mongo.db.rental_requests.insert_one({
            'item_id': item_id,
            'user_name': user_name,
            'user_email': user_email,
            'start_date': datetime.strptime(start_date, '%Y-%m-%d'),
            'end_date': datetime.strptime(end_date, '%Y-%m-%d'),
            'quantity': quantity,
            'purpose': purpose,
            'status': 'Pending',
            'created_at': datetime.now()
        })
        
        flash('Rental request submitted successfully! Staff will review your request.', 'success')
        return redirect(url_for('index'))
    
    # GET request - show rental form
    item = mongo.db.cars.find_one({'id': item_id})
    if not item:
        flash('Voiture non trouvée', 'error')
        return redirect(url_for('index'))
    
    # Ensure the id field is properly set for the template
    item['id'] = item.get('id', item_id)
    
    return render_template('rental_request.html', item=item)

# Update staff_requests route
@app.route('/staff/requests')
@login_required
def staff_requests():
    # Only manager and admin can access
    if not (is_manager() or current_user.role == 'admin'):
        flash('Accès refusé. Réservé au manager et admin.', 'error')
        return redirect(url_for('index'))
    
    # Get all rental requests
    requests = list(mongo.db.rental_requests.find().sort('created_at', -1))
    
    for reservation in requests:
        reservation['id'] = str(reservation['_id'])
        reservation['created_at'] = reservation.get('created_at', '').strftime('%Y-%m-%d %H:%M') if reservation.get('created_at') else ''
        
        # Format dates for template display
        if reservation.get('start_date'):
            if isinstance(reservation['start_date'], datetime):
                reservation['start_date'] = reservation['start_date'].strftime('%Y-%m-%d')
            else:
                reservation['start_date'] = str(reservation['start_date'])
        else:
            reservation['start_date'] = 'N/A'
            
        if reservation.get('end_date'):
            if isinstance(reservation['end_date'], datetime):
                reservation['end_date'] = reservation['end_date'].strftime('%Y-%m-%d')
            else:
                reservation['end_date'] = str(reservation['end_date'])
        else:
            reservation['end_date'] = 'N/A'
        
        # Handle car name and quantity
        if 'items' in reservation and reservation['items'] and isinstance(reservation['items'], list):
            # Multi-item request
            print(f"DEBUG: Multi-item reservation: {reservation['items']}")
            item_names = []
            total_quantity = 0
            for item_data in reservation['items']:
                if isinstance(item_data, dict):
                    car = mongo.db.cars.find_one({'id': item_data.get('item_id')})
                    if car:
                        item_names.append(f"{car.get('designation', 'Unknown')} (x{item_data.get('quantity', 1)})")
                        total_quantity += item_data.get('quantity', 1)
                    else:
                        item_names.append(f"Unknown (x{item_data.get('quantity', 1)})")
                        total_quantity += item_data.get('quantity', 1)
            
            reservation['car_name'] = ' + '.join(item_names) if item_names else 'Unknown'
            reservation['quantity'] = total_quantity
        else:
            # Single-item request (legacy)
            print(f"DEBUG: Single-item reservation: {reservation.get('item_id')}")
            car = mongo.db.cars.find_one({'id': reservation.get('item_id')})
            reservation['car_name'] = car['designation'] if car else 'Unknown'
            reservation['quantity'] = reservation.get('quantity', 1)
    
    # Get all users for role checking
    users = list(mongo.db.users.find({}, {'username': 1, 'role': 1}))
    
    return render_template('staff_requests.html', requests=requests, users=users)

# Update approve_request route
@app.route('/staff/approve-request/<string:req_id>')
@login_required
def approve_request(req_id):
    # Manager and admin can approve requests
    if not (is_manager() or current_user.role == 'admin'):
        flash('Accès refusé.', 'error')
        return redirect(url_for('index'))
    
    request_obj = mongo.db.rental_requests.find_one({'_id': ObjectId(req_id)})
    if request_obj:
        # Handle both single-item and multi-item requests
        if 'items' in request_obj:
            # Multi-item request - quantities were already deducted when created
            # Just update the status to approved
            mongo.db.rental_requests.update_one(
                {'_id': ObjectId(req_id)},
                {'$set': {'status': 'Approved'}}
            )
            flash('Multi-item request approved successfully!', 'success')
        else:
            # Single-item request (legacy)
            car = mongo.db.cars.find_one({'id': request_obj.get('item_id')})
            if not car:
                flash('Voiture non trouvée', 'error')
                return redirect(url_for('staff_requests'))
            
            requested_quantity = request_obj.get('quantity', 1)
            current_available = car.get('quantite_disponible', car.get('quantite_totale', 1))
            
            # Check if enough quantity is available
            if requested_quantity > current_available:
                flash(f'Pas assez d\'unités disponibles. Demandé: {requested_quantity}, Disponible: {current_available}', 'error')
                return redirect(url_for('staff_requests'))
            
            # Calculate new available quantity
            new_available = current_available - requested_quantity
            
            # Update request status
            mongo.db.rental_requests.update_one(
                {'_id': ObjectId(req_id)},
                {'$set': {'status': 'Approved'}}
            )
            
            # Update car quantity and status
            update_fields = {
                'quantite_disponible': new_available,
                'updated_at': datetime.now()
            }
            
            # If no units left, mark as unavailable
            if new_available <= 0:
                update_fields['status'] = 'Indisponible'
            
            mongo.db.cars.update_one(
                {'id': request_obj.get('item_id')},
                {'$set': update_fields}
            )
            
            flash(f'Request approved successfully! {requested_quantity} units rented.', 'success')
    else:
        flash('Request not found', 'error')
    
    return redirect(url_for('staff_requests'))

# Update reject_request route
@app.route('/staff/reject-request/<string:req_id>')
@login_required
def reject_request(req_id):
    # Manager and admin can reject requests
    if not (is_manager() or current_user.role == 'admin'):
        flash('Accès refusé.', 'error')
        return redirect(url_for('index'))
    
    request_obj = mongo.db.rental_requests.find_one({'_id': ObjectId(req_id)})
    if request_obj:
        
        # Handle both single-item and multi-item requests
        if 'items' in request_obj:
            # Multi-item request - restore quantities for each item
            for item_data in request_obj['items']:
                car = mongo.db.cars.find_one({'id': item_data.get('item_id')})
                if car:
                    current_available = car.get('quantite_disponible', 0)
                    requested_quantity = item_data.get('quantity', 1)
                    new_available = current_available + requested_quantity
                    
                    # Update car quantity
                    mongo.db.cars.update_one(
                        {'id': item_data.get('item_id')},
                        {'$set': {
                            'quantite_disponible': new_available,
                            'status': 'Disponible' if new_available > 0 else 'Indisponible',
                            'updated_at': datetime.now()
                        }}
                    )
        else:
            # Single-item request - restore quantity
            car = mongo.db.cars.find_one({'id': request_obj.get('item_id')})
            if car:
                current_available = car.get('quantite_disponible', 0)
                requested_quantity = request_obj.get('quantity', 1)
                new_available = current_available + requested_quantity
                
                # Update car quantity
                mongo.db.cars.update_one(
                    {'id': request_obj.get('item_id')},
                    {'$set': {
                        'quantite_disponible': new_available,
                        'status': 'Disponible' if new_available > 0 else 'Indisponible',
                        'updated_at': datetime.now()
                    }}
                )
    
    mongo.db.rental_requests.update_one(
        {'_id': ObjectId(req_id)},
        {'$set': {'status': 'Rejected'}}
    )
    flash('Request rejected successfully!', 'success')
    return redirect(url_for('staff_requests'))

# Add reset request status route
@app.route('/staff/reset-request-status/<string:req_id>')
@login_required
def reset_request_status(req_id):
    # Only manager and admin can reset status
    if not (is_manager() or current_user.role == 'admin'):
        flash('Accès refusé.', 'error')
        return redirect(url_for('index'))
    
    mongo.db.rental_requests.update_one(
        {'_id': ObjectId(req_id)},
        {'$set': {'status': 'En attente'}}
    )
    flash('Request status reset to pending', 'success')
    return redirect(url_for('staff_requests'))

@app.route('/staff/cars-used')
@login_required
def staff_cars_used():
    """Staff view of currently rented cars"""
    print(f"DEBUG: staff_cars_used called by user: {current_user.username}, role: {current_user.role}")
    
    if not (is_manager() or current_user.role == 'admin'):
        print(f"DEBUG: Access denied for user {current_user.username} with role {current_user.role}")
        flash('Accès refusé', 'error')
        return redirect(url_for('dashboard'))
    
    print(f"DEBUG: Access granted, proceeding to render template")
    
    # Get only APPROVED reservations that are currently in use (not completed, returned, rejected, or pending)
    active_reservations = list(mongo.db.rental_requests.find({
        'status': {'$in': ['Approved', 'Active']}  # Only show reservations that were actually approved and given to users
    }).sort('start_date', -1))
    
    # Format reservations for template
    formatted_reservations = []
    for r in active_reservations:
        if 'items' in r and r['items']:
            # Multi-item reservation
            item_names = []
            total_quantity = 0
            for item_data in r['items']:
                item_id = item_data.get('item_id', '')
                quantity = item_data.get('quantity', 1)
                designation = item_data.get('designation', '')
                item_names.append(f"{designation} (x{quantity})")
                total_quantity += quantity
            
            formatted_reservations.append({
                'id': str(r.get('_id')),
                'item_name': ' + '.join(item_names),
                'user_name': r.get('user_name', ''),
                'user_email': r.get('user_email', ''),
                'quantity': total_quantity,
                'start_date': r.get('start_date', '').strftime('%Y-%m-%d %H:%M') if r.get('start_date') else '',
                'end_date': r.get('end_date', '').strftime('%Y-%m-%d %H:%M') if r.get('end_date') else '',
                'purpose': r.get('purpose', ''),
                'is_multi_item': True
            })
        else:
            # Single item reservation (legacy)
            car = mongo.db.cars.find_one({'id': r.get('item_id')}) or {}
            formatted_reservations.append({
                'id': str(r.get('_id')),
                'item_name': car.get('designation', ''),
                'user_name': r.get('user_name', ''),
                'user_email': r.get('user_email', ''),
                'quantity': r.get('quantity', 1),
                'start_date': r.get('start_date', '').strftime('%Y-%m-%d %H:%M') if r.get('start_date') else '',
                'end_date': r.get('end_date', '').strftime('%Y-%m-%d %H:%M') if r.get('end_date') else '',
                'purpose': r.get('purpose', ''),
                'is_multi_item': False
            })
    
    return render_template('staff_rented_cars.html', reservations=formatted_reservations)

@app.route('/staff/return-car/<string:item_id>', methods=['GET', 'POST'])
@login_required
def return_car(item_id):
    if not (is_manager() or current_user.role == 'admin'):
        flash('Accès refusé. Réservé au manager ou admin.', 'error')
        return redirect(url_for('index'))
    
    # Get the rental request to know how many units were rented
    # Check both single-item and multi-item requests
    rental_request = mongo.db.rental_requests.find_one({
        '$or': [
            {'item_id': item_id, 'status': 'Approved'},
            {'items.item_id': item_id, 'status': 'Approved'}
        ]
    })
    
    if not rental_request:
        flash('Aucune utilisation active trouvée pour cet équipement', 'error')
        return redirect(url_for('staff_cars_used'))
    
    # Get current car
    car = mongo.db.cars.find_one({'id': item_id})
    if not car:
        flash('Équipement non trouvé', 'error')
        return redirect(url_for('staff_cars_used'))
    
    # Determine rented quantity based on request type
    if 'items' in rental_request:
        # Multi-item request - find the specific item
        item_data = next((item for item in rental_request['items'] if item['item_id'] == item_id), None)
        rented_quantity = item_data['quantity'] if item_data else 1
    else:
        # Single-item request
        rented_quantity = rental_request.get('quantity', 1)
    current_available = car.get('quantite_disponible', car.get('quantite_totale', 1))
    
    if request.method == 'GET':
        # Show the return form
        return render_template('return_car.html', 
                             car=car, 
                             rental_request=rental_request,
                             rented_quantity=rented_quantity,
                             current_available=current_available)
    
    elif request.method == 'POST':
        # Process the return
        new_status = request.form.get('new_status')
        notes = request.form.get('notes', '')
        
        if not new_status:
            flash('Veuillez sélectionner un statut', 'error')
            return redirect(url_for('return_car', item_id=item_id))
        
        # Calculate new quantities based on status
        quantite_disponible = current_available + rented_quantity
        quantite_cassee = car.get('quantite_cassée', 0)
        quantite_en_reparation = car.get('quantite_en_réparation', 0)
        
        # Adjust quantities based on selected status
        if new_status == 'Cassée':
            quantite_cassee += rented_quantity
            quantite_disponible -= rented_quantity
        elif new_status == 'En réparation':
            quantite_en_reparation += rented_quantity
            quantite_disponible -= rented_quantity
        # For 'Disponible' and 'Indisponible', quantities stay as calculated above
        
        # Update car
        update_fields = {
            'quantite_disponible': quantite_disponible,
            'quantite_cassée': quantite_cassee,
            'quantite_en_réparation': quantite_en_reparation,
            'status': new_status,
            'updated_at': datetime.now()
        }
        
        if notes:
            update_fields['notes'] = notes
        
        mongo.db.cars.update_one(
            {'id': item_id},
            {'$set': update_fields}
        )
        
        # Update the rental request status to Completed
        # Handle both single-item and multi-item requests
        if 'items' in rental_request:
            # Multi-item request - update the specific item status
            mongo.db.rental_requests.update_one(
                {'_id': rental_request['_id']},
                {'$set': {'items.$.status': 'Completed'}}
            )
        else:
            # Single-item request
            mongo.db.rental_requests.update_one(
                {'item_id': item_id, 'status': 'Approved'},
                {'$set': {'status': 'Completed'}}
            )
        
        flash(f'Équipement retourné avec succès! {rented_quantity} unités marquées comme "{new_status}".', 'success')
        return redirect(url_for('staff_cars_used'))

@app.route('/staff/edit-car/<string:item_id>', methods=['GET', 'POST'])
@login_required
def edit_car(item_id):
    if not (is_manager() or current_user.role == 'admin'):
        flash('Accès refusé. Réservé au manager ou admin.', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        id_car = request.form.get('id')
        quantity = int(request.form.get('quantity', 1))
        prix_journalier = float(request.form.get('prix_journalier', 0)) if request.form.get('prix_journalier') else 0
        carburant = request.form.get('carburant', 'Essence')
        transmission = request.form.get('transmission', 'Manuelle')
        designation = request.form.get('designation')
        marque = request.form.get('marque')
        modele = request.form.get('modele')
        n_serie = request.form.get('n_serie')
        ancien_cab = request.form.get('ancien_cab')
        nouveau_cab = request.form.get('nouveau_cab')
        status = request.form.get('status')
        date_inv = request.form.get('date_inv')
        description = request.form.get('description')
        # Keep existing quantity values from database
        existing_item = mongo.db.cars.find_one({'id': item_id})
        quantite_totale = existing_item.get('quantite_totale', 1)
        quantite_cassee = existing_item.get('quantite_cassée', 0)
        quantite_en_reparation = existing_item.get('quantite_en_réparation', 0)
        quantite_disponible = existing_item.get('quantite_disponible', quantite_totale - quantite_cassee - quantite_en_reparation)
        image_file = request.files.get('image')
        image_filename = None
        
        if image_file and image_file.filename:
            if not allowed_file(image_file.filename):
                flash('Type de fichier invalide. Seuls JPG, PNG, GIF autorisés.', 'danger')
                return redirect(request.url)
            if len(image_file.read()) > app.config['MAX_CONTENT_LENGTH']:
                flash('Le fichier image est trop volumineux (max 2MB).', 'danger')
                return redirect(request.url)
            image_file.seek(0)
            filename = secure_filename(image_file.filename)
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            image_file.save(image_path)
            image_filename = filename
        
        update_fields = {
            'id': id_car,
            'quantity': quantity,
            'prix_journalier': prix_journalier,
            'carburant': carburant,
            'transmission': transmission,
            'designation': designation,
            'marque': marque,
            'modele': modele,
            'n_serie': n_serie,
            'ancien_cab': ancien_cab,
            'nouveau_cab': nouveau_cab,
            'status': status,
            'date_inv': date_inv,
            'description': description,
            'quantite_totale': quantite_totale,
            'quantite_cassée': quantite_cassee,
            'quantite_en_réparation': quantite_en_reparation,
            'quantite_disponible': quantite_disponible,
            'updated_at': datetime.now()
        }
        
        if image_filename:
            update_fields['image'] = image_filename
        
        mongo.db.cars.update_one(
            {'id': item_id},
            {'$set': update_fields}
        )
        
        flash('Voiture mise à jour avec succès!', 'success')
        return redirect(url_for('inventory'))
    
    # GET request - show edit form
    item = mongo.db.cars.find_one({'id': item_id})
    if not item:
        flash('Voiture non trouvée', 'error')
        return redirect(url_for('inventory'))
    
    # Don't overwrite the actual 'id' field with MongoDB _id
    # The 'id' field should remain as the car's custom ID
    return render_template('edit_car.html', item=item)

@app.route('/view-car/<string:item_id>')
@login_required
def view_car(item_id):
    item = mongo.db.cars.find_one({'id': item_id})
    if not item:
        flash('Voiture non trouvée', 'error')
        return redirect(url_for('inventory'))
    
    # Don't overwrite the actual 'id' field with MongoDB _id
    item['created_at'] = item.get('created_at', '').strftime('%Y-%m-%d %H:%M') if item.get('created_at') else ''
    item['updated_at'] = item.get('updated_at', '').strftime('%Y-%m-%d %H:%M') if item.get('updated_at') else ''
    
    # Get rental history for this car
    # Search for both single-item and multi-item requests
    single_item_rentals = list(mongo.db.rental_requests.find({'item_id': item_id}).sort('created_at', -1))
    multi_item_rentals = list(mongo.db.rental_requests.find({'items.item_id': item_id}).sort('created_at', -1))
    
    # Combine and process all rentals
    rental_history = []
    
    # Process single-item rentals
    for rental in single_item_rentals:
        rental['id'] = str(rental['_id'])
        rental['created_at'] = rental.get('created_at', '').strftime('%Y-%m-%d %H:%M') if rental.get('created_at') else ''
        rental['start_date'] = rental.get('start_date', '').strftime('%Y-%m-%d') if rental.get('start_date') else ''
        rental['end_date'] = rental.get('end_date', '').strftime('%Y-%m-%d') if rental.get('end_date') else ''
        rental['is_multi_item'] = False
        rental_history.append(rental)
    
    # Process multi-item rentals
    for rental in multi_item_rentals:
        # Find the specific item in the multi-item request
        item_data = next((item for item in rental.get('items', []) if item.get('item_id') == item_id), None)
        if item_data:
            rental['id'] = str(rental['_id'])
            rental['created_at'] = rental.get('created_at', '').strftime('%Y-%m-%d %H:%M') if rental.get('created_at') else ''
            rental['start_date'] = rental.get('start_date', '').strftime('%Y-%m-%d') if rental.get('start_date') else ''
            rental['end_date'] = rental.get('end_date', '').strftime('%Y-%m-%d') if rental.get('end_date') else ''
            rental['is_multi_item'] = True
            rental['item_quantity'] = item_data.get('quantity', 1)
            rental_history.append(rental)
    
    # Sort by creation date
    rental_history.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    
    return render_template('view_car.html', item=item, rental_history=rental_history)

@app.route('/staff/delete-car/<string:item_id>')
@login_required
def delete_car(item_id):
    if not (is_manager() or current_user.role == 'admin'):
        flash('Accès refusé. Réservé au manager ou admin.', 'error')
        return redirect(url_for('index'))
    
    # Check if car is currently rented
    item = mongo.db.cars.find_one({'id': item_id})
    if item and item.get('status') == 'Indisponible':
        flash('Impossible de supprimer une voiture actuellement louée.', 'error')
        return redirect(url_for('inventory'))
    
    # Delete the car
    mongo.db.cars.delete_one({'id': item_id})
    
    # Also delete any related rental requests
    # Delete single-item requests
    mongo.db.rental_requests.delete_many({'item_id': item_id})
    # Delete multi-item requests that contain this item
    mongo.db.rental_requests.delete_many({'items.item_id': item_id})
    
    flash('Voiture supprimée avec succès!', 'success')
    return redirect(url_for('inventory'))

# Staff Management Routes (Admin Only)
@app.route('/admin/staff')
@login_required
def admin_staff():
    if current_user.role != 'admin':
        flash('Access denied. Admin only.', 'error')
        return redirect(url_for('index'))
    
    # Show all users except admin
    staff_members = list(mongo.db.users.find({'role': {'$ne': 'admin'}}).sort('username', 1))
    for staff in staff_members:
        staff['id'] = str(staff['_id'])
        staff['created_at'] = staff.get('created_at', '').strftime('%Y-%m-%d %H:%M') if staff.get('created_at') else ''
    
    return render_template('admin_staff.html', staff_members=staff_members)

@app.route('/admin/add-staff', methods=['GET', 'POST'])
@login_required
def add_staff():
    if current_user.role != 'admin':
        flash('Access denied. Admin only.', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        role = request.form.get('role')
        
        if not all([username, password, first_name, last_name]):
            flash('Please fill in all required fields', 'error')
            return redirect(url_for('add_staff'))

        if password != confirm_password:
            flash('Les mots de passe ne correspondent pas', 'error')
            return redirect(url_for('add_staff'))
        
        # Check if username already exists
        if mongo.db.users.find_one({'username': username}):
            flash('Username already exists', 'error')
            return redirect(url_for('add_staff'))
        
        # Create new user with selected role
        hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
        mongo.db.users.insert_one({
            'username': username,
            'password': hashed_password,
            'role': role,
            'first_name': first_name,
            'last_name': last_name,
            'created_at': datetime.now()
        })
        
        flash(f'User {first_name} {last_name} added successfully!', 'success')
        return redirect(url_for('admin_staff'))
    
    return render_template('add_staff.html')

@app.route('/admin/edit-staff/<string:staff_id>', methods=['GET', 'POST'])
@login_required
def edit_staff(staff_id):
    if current_user.role != 'admin':
        flash('Access denied. Admin only.', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if not all([username, first_name, last_name]):
            flash('Please fill in all required fields', 'error')
            return redirect(url_for('edit_staff', staff_id=staff_id))
        
        # Check if username already exists (excluding current user)
        existing_user = mongo.db.users.find_one({'username': username, '_id': {'$ne': ObjectId(staff_id)}})
        if existing_user:
            flash('Username already exists', 'error')
            return redirect(url_for('edit_staff', staff_id=staff_id))
        
        # Update staff member
        update_fields = {
            'username': username,
            'first_name': first_name,
            'last_name': last_name
        }
        
        if password:
            if password != confirm_password:
                flash('Les mots de passe ne correspondent pas', 'error')
                return redirect(url_for('edit_staff', staff_id=staff_id))
            update_fields['password'] = bcrypt.generate_password_hash(password).decode('utf-8')
        
        mongo.db.users.update_one(
            {'_id': ObjectId(staff_id)},
            {'$set': update_fields}
        )
        
        flash('Staff member updated successfully!', 'success')
        return redirect(url_for('admin_staff'))
    
    # GET request - show edit form (any non-admin user)
    staff = mongo.db.users.find_one({'_id': ObjectId(staff_id), 'role': {'$ne': 'admin'}})
    if not staff:
        flash('Membre du personnel introuvable', 'error')
        return redirect(url_for('admin_staff'))
    
    staff['id'] = str(staff['_id'])
    return render_template('edit_staff.html', staff=staff)

@app.route('/admin/delete-staff/<string:staff_id>')
@login_required
def delete_staff(staff_id):
    if current_user.role != 'admin':
        flash('Access denied. Admin only.', 'error')
        return redirect(url_for('index'))
    
    # Don't allow admin to delete themselves
    if str(current_user.id) == staff_id:
        flash('You cannot delete your own account', 'error')
        return redirect(url_for('admin_staff'))
    
    mongo.db.users.delete_one({'_id': ObjectId(staff_id), 'role': {'$ne': 'admin'}})
    flash('User deleted successfully!', 'success')
    return redirect(url_for('admin_staff'))

# API endpoint for staff requests data
@app.route('/api/staff/requests')
@login_required
def api_staff_requests():
    # Only manager and admin can access
    if not (is_manager() or current_user.role == 'admin'):
        return jsonify({'error': 'Accès refusé'}), 403
    
    # Get all rental requests
    requests = list(mongo.db.rental_requests.find().sort('created_at', -1))
    
    # Process requests for JSON response
    processed_requests = []
    for reservation in requests:
        reservation['id'] = str(reservation['_id'])
        
        # Handle both single-item and multi-item requests
        if 'items' in reservation and reservation['items']:
            # Multi-item request
            item_names = []
            for item_data in reservation['items']:
                if isinstance(item_data, dict):
                    car = mongo.db.cars.find_one({'id': item_data.get('item_id')})
                    if car:
                        item_names.append(f"{car.get('designation', 'Unknown')} (x{item_data.get('quantity', 1)})")
                    else:
                        item_names.append(f"Unknown (x{item_data.get('quantity', 1)})")
            reservation['car_name'] = ' + '.join(item_names) if item_names else 'Unknown'
        else:
            # Single-item request (legacy)
            car = mongo.db.cars.find_one({'id': reservation.get('item_id')})
            reservation['car_name'] = car['designation'] if car else 'Unknown'
        
        # Convert datetime objects to strings for JSON
        if reservation.get('start_date'):
            if isinstance(reservation['start_date'], datetime):
                reservation['start_date'] = reservation['start_date'].isoformat()
        if reservation.get('end_date'):
            if isinstance(reservation['end_date'], datetime):
                reservation['end_date'] = reservation['end_date'].isoformat()
        if reservation.get('created_at'):
            if isinstance(reservation['created_at'], datetime):
                reservation['created_at'] = reservation['created_at'].isoformat()
        
        processed_requests.append(reservation)
    
    return jsonify({'requests': processed_requests})

# Route to mark car as returned (AJAX)
@app.route('/staff/mark-returned/<string:reservation_id>', methods=['POST'])
@login_required
def mark_car_returned(reservation_id):
    """Mark car as returned via AJAX"""
    if not (is_manager() or current_user.role == 'admin'):
        return jsonify({'success': False, 'message': 'Accès refusé'}), 403
    
    try:
        # Validate ObjectId format
        try:
            object_id = ObjectId(reservation_id)
        except Exception as e:
            print(f"Invalid ObjectId format: {reservation_id}")
            return jsonify({'success': False, 'message': f'ID de réservation invalide: {reservation_id}'}), 400
        
        data = request.get_json()
        if not data:
            print("No JSON data received")
            return jsonify({'success': False, 'message': 'Aucune donnée reçue'}), 400
            
        action = data.get('action')
        print(f"Action received: {action}")
        
        if action == 'mark_returned':
            # Find the reservation
            reservation = mongo.db.rental_requests.find_one({'_id': object_id})
            if not reservation:
                print(f"Reservation not found: {reservation_id}")
                return jsonify({'success': False, 'message': 'Réservation non trouvée'}), 404
            
            print(f"Found reservation: {reservation}")
            
            # Get status selections from request
            status_selections = data.get('status_selections', [])
            if not status_selections:
                return jsonify({'success': False, 'message': 'Aucune sélection de statut fournie'}), 400
            
            print(f"Status selections: {status_selections}")
            
            # Handle both single-item and multi-item reservations
            if 'items' in reservation and reservation['items']:
                # Multi-item reservation
                print("Processing multi-item reservation")
                for item_data in reservation['items']:
                    # Find corresponding status selection
                    status_selection = next((s for s in status_selections if s['item_id'] == item_data.get('item_id')), None)
                    if not status_selection:
                        continue
                    
                    car = mongo.db.cars.find_one({'id': item_data.get('item_id')})
                    if car:
                        # Get current quantities
                        current_available = car.get('quantite_disponible', 0)
                        current_broken = car.get('quantite_cassée', 0)
                        current_repair = car.get('quantite_en_réparation', 0)
                        current_unavailable = car.get('quantite_indisponible', 0)
                        current_lost = car.get('quantite_perdue', 0)
                        
                        returned_quantity = item_data.get('quantity', 1)
                        new_status = status_selection['status']
                        
                        print(f"Updating car {item_data.get('item_id')}: status={new_status}, quantity={returned_quantity}")
                        
                        # Update quantities based on selected status
                        update_fields = {
                            'updated_at': datetime.now()
                        }
                        
                        if new_status == 'Disponible':
                            update_fields['quantite_disponible'] = current_available + returned_quantity
                            update_fields['status'] = 'Disponible'
                        elif new_status == 'Cassée':
                            update_fields['quantite_cassée'] = current_broken + returned_quantity
                            update_fields['status'] = 'Cassée'
                        elif new_status == 'En réparation':
                            update_fields['quantite_en_réparation'] = current_repair + returned_quantity
                            update_fields['status'] = 'En réparation'
                        elif new_status == 'Indisponible':
                            update_fields['quantite_indisponible'] = current_unavailable + returned_quantity
                            update_fields['status'] = 'Indisponible'
                        elif new_status == 'Perdue':
                            update_fields['quantite_perdue'] = current_lost + returned_quantity
                            update_fields['status'] = 'Perdue'
                        
                        # Update car
                        mongo.db.cars.update_one(
                            {'id': item_data.get('item_id')},
                            {'$set': update_fields}
                        )
            else:
                # Single-item reservation (legacy)
                print("Processing single-item reservation")
                if len(status_selections) > 0:
                    status_selection = status_selections[0]
                    car = mongo.db.cars.find_one({'id': reservation.get('item_id')})
                    if car:
                        # Get current quantities
                        current_available = car.get('quantite_disponible', 0)
                        current_broken = car.get('quantite_cassée', 0)
                        current_repair = car.get('quantite_en_réparation', 0)
                        current_unavailable = car.get('quantite_indisponible', 0)
                        current_lost = car.get('quantite_perdue', 0)
                        
                        returned_quantity = reservation.get('quantity', 1)
                        new_status = status_selection['status']
                        
                        print(f"Updating car {reservation.get('item_id')}: status={new_status}, quantity={returned_quantity}")
                        
                        # Update quantities based on selected status
                        update_fields = {
                            'updated_at': datetime.now()
                        }
                        
                        if new_status == 'Disponible':
                            update_fields['quantite_disponible'] = current_available + returned_quantity
                            update_fields['status'] = 'Disponible'
                        elif new_status == 'Cassée':
                            update_fields['quantite_cassée'] = current_broken + returned_quantity
                            update_fields['status'] = 'Cassée'
                        elif new_status == 'En réparation':
                            update_fields['quantite_en_réparation'] = current_repair + returned_quantity
                            update_fields['status'] = 'En réparation'
                        elif new_status == 'Indisponible':
                            update_fields['quantite_indisponible'] = current_unavailable + returned_quantity
                            update_fields['status'] = 'Indisponible'
                        elif new_status == 'Perdue':
                            update_fields['quantite_perdue'] = current_lost + returned_quantity
                            update_fields['status'] = 'Perdue'
                        
                        # Update car
                        mongo.db.cars.update_one(
                            {'id': reservation.get('item_id')},
                            {'$set': update_fields}
                        )
            
            # Update reservation status to 'Completed'
            mongo.db.rental_requests.update_one(
                {'_id': object_id},
                {'$set': {'status': 'Completed', 'returned_at': datetime.now()}}
            )
            
            print("Successfully marked as returned")
            return jsonify({'success': True, 'message': 'Matériel marqué comme retourné avec succès'})
        else:
            print(f"Unknown action: {action}")
            return jsonify({'success': False, 'message': 'Action non reconnue'}), 400
            
    except Exception as e:
        print(f"Error marking car as returned: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Erreur lors du marquage: {str(e)}'}), 500

# Shutdown endpoint for the launcher
@app.route('/shutdown', methods=['POST'])
def shutdown():
    """Shutdown the Flask application"""
    import os
    os._exit(0)

if __name__ == '__main__':
    # Get port from environment variable (for Netlify)
    port = int(os.environ.get('PORT', 5000))
    # Run in production mode for deployment
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug_mode)

# Add URL rules for the staff request routes
app.add_url_rule('/staff/approve-request/<string:req_id>', 'approve_request', approve_request)
app.add_url_rule('/staff/reject-request/<string:req_id>', 'reject_request', reject_request)
app.add_url_rule('/staff/reset-request-status/<string:req_id>', 'reset_request_status', reset_request_status)
